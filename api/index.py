from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
import os
import tempfile
from datetime import datetime, timedelta
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import uuid

app = Flask(__name__, 
           template_folder='../templates',
           static_folder='../static')

# Configuration
app.secret_key = os.environ.get('SECRET_KEY', 'your-secret-key-here-change-in-production')
app.permanent_session_lifetime = timedelta(hours=1)

# Admin credentials
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'pahlawan140')

# Global variables for file paths
EXCEL_DIR = tempfile.gettempdir()
EXCEL_FILE = os.path.join(EXCEL_DIR, 'data_sensus.xlsx')

def create_excel_file():
    """Create Excel file with proper headers and formatting"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Sensus"
        
        # Headers
        headers = [
            'ID Keluarga', 'RT', 'RW', 'Dusun', 'Nama Kepala Keluarga', 'Alamat',
            'Jumlah Anggota Keluarga', 'Jumlah Anggota Usia 15+',
            'Nama Anggota', 'Umur', 'Hubungan dengan Kepala Keluarga', 'Jenis Kelamin',
            'Status Perkawinan', 'Pendidikan Terakhir', 'Kegiatan Sehari-hari',
            'Apakah Memiliki Pekerjaan', 'Status Pekerjaan yang Diinginkan',
            'Bidang Usaha yang Diminati', 'Bidang Pekerjaan', 'Status Pekerjaan Utama',
            'Pemasaran Usaha Utama', 'Penjualan Marketplace Utama',
            'Memiliki Lebih dari Satu Pekerjaan', 'Bidang Pekerjaan Sampingan 1',
            'Status Pekerjaan Sampingan 1', 'Pemasaran Usaha Sampingan 1',
            'Penjualan Marketplace Sampingan 1', 'Bidang Pekerjaan Sampingan 2',
            'Status Pekerjaan Sampingan 2', 'Pemasaran Usaha Sampingan 2',
            'Penjualan Marketplace Sampingan 2', 'Nama Pencacah', 'HP Pencacah',
            'Tanggal Pencacah', 'Nama Pemberi Jawaban', 'HP Pemberi Jawaban',
            'Tanggal Pemberi Jawaban', 'Catatan'
        ]
        
        # Add headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Save the workbook
        wb.save(EXCEL_FILE)
        print(f"Excel file created at: {EXCEL_FILE}")
        return True
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return False

def save_to_excel(keluarga_data, all_members_data, final_data):
    """Save all data to Excel file"""
    try:
        # Create file if it doesn't exist
        if not os.path.exists(EXCEL_FILE):
            create_excel_file()
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Prepare data for each member
        for i, member_data in enumerate(all_members_data):
            row_data = [
                keluarga_data.get('keluarga_id', ''),
                keluarga_data.get('rt', ''),
                keluarga_data.get('rw', ''),
                keluarga_data.get('dusun', ''),
                keluarga_data.get('nama_kepala', ''),
                keluarga_data.get('alamat', ''),
                keluarga_data.get('jumlah_anggota', ''),
                keluarga_data.get('jumlah_anggota_15plus', ''),
                member_data.get('Nama Anggota', ''),
                member_data.get('Umur', ''),
                member_data.get('Hubungan dengan Kepala Keluarga', ''),
                member_data.get('Jenis Kelamin', ''),
                member_data.get('Status Perkawinan', ''),
                member_data.get('Pendidikan Terakhir', ''),
                member_data.get('Kegiatan Sehari-hari', ''),
                member_data.get('Apakah Memiliki Pekerjaan', ''),
                member_data.get('Status Pekerjaan yang Diinginkan', ''),
                member_data.get('Bidang Usaha yang Diminati', ''),
                member_data.get('Bidang Pekerjaan', ''),
                member_data.get('Status Pekerjaan Utama', ''),
                member_data.get('Pemasaran Usaha Utama', ''),
                member_data.get('Penjualan Marketplace Utama', ''),
                member_data.get('Memiliki Lebih dari Satu Pekerjaan', ''),
                member_data.get('Bidang Pekerjaan Sampingan 1', ''),
                member_data.get('Status Pekerjaan Sampingan 1', ''),
                member_data.get('Pemasaran Usaha Sampingan 1', ''),
                member_data.get('Penjualan Marketplace Sampingan 1', ''),
                member_data.get('Bidang Pekerjaan Sampingan 2', ''),
                member_data.get('Status Pekerjaan Sampingan 2', ''),
                member_data.get('Pemasaran Usaha Sampingan 2', ''),
                member_data.get('Penjualan Marketplace Sampingan 2', ''),
                final_data.get('nama_pencacah', '') if i == 0 else '',
                final_data.get('hp_pencacah', '') if i == 0 else '',
                final_data.get('tanggal_pencacah', '') if i == 0 else '',
                final_data.get('nama_pemberi_jawaban', '') if i == 0 else '',
                final_data.get('hp_pemberi_jawaban', '') if i == 0 else '',
                final_data.get('tanggal_pemberi_jawaban', '') if i == 0 else '',
                final_data.get('catatan', '') if i == 0 else ''
            ]
            
            # Add data to worksheet
            for col, value in enumerate(row_data, 1):
                ws.cell(row=next_row, column=col, value=value)
            
            next_row += 1
        
        # Save the workbook
        wb.save(EXCEL_FILE)
        print(f"Data saved to Excel file: {EXCEL_FILE}")
        return True
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return False

# Routes
@app.route('/')
def login():
    if 'logged_in' in session and session['logged_in']:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def handle_login():
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session.permanent = True
            session['logged_in'] = True
            session['username'] = username
            session['login_time'] = datetime.now().isoformat()
            return jsonify({
                'success': True,
                'message': 'Login berhasil',
                'redirect_url': '/dashboard'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Username atau password salah!'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': 'Terjadi kesalahan pada server'
        })

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/check-session')
def check_session():
    if 'logged_in' in session and session['logged_in']:
        # Check if session has expired
        if 'login_time' in session:
            login_time = datetime.fromisoformat(session['login_time'])
            if datetime.now() - login_time > timedelta(hours=1):
                session.clear()
                return jsonify({'logged_in': False, 'expired': True})
        return jsonify({'logged_in': True})
    return jsonify({'logged_in': False})

def login_required(f):
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session or not session['logged_in']:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/index')
@login_required
def index():
    return render_template('index.html')

@app.route('/get-form-data')
@login_required
def get_form_data():
    form_data = session.get('form_data', {})
    return jsonify({'form_data': form_data})

@app.route('/submit', methods=['POST'])
@login_required
def submit():
    try:
        # Generate unique family ID
        keluarga_id = str(uuid.uuid4())[:8].upper()
        
        # Get form data
        form_data = {
            'keluarga_id': keluarga_id,
            'rt': request.form.get('rt'),
            'rw': request.form.get('rw'),
            'dusun': request.form.get('dusun'),
            'nama_kepala': request.form.get('nama_kepala'),
            'alamat': request.form.get('alamat'),
            'jumlah_anggota': int(request.form.get('jumlah_anggota', 0)),
            'jumlah_anggota_15plus': int(request.form.get('jumlah_anggota_15plus', 0))
        }
        
        # Store in session
        session['keluarga_data'] = form_data
        session['current_member'] = 0
        session['all_members_data'] = []
        
        # Clear any previous form data
        if 'form_data' in session:
            del session['form_data']
        
        return jsonify({
            'success': True,
            'message': 'Data keluarga berhasil disimpan',
            'redirect': True,
            'redirect_url': '/lanjutan'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        })

@app.route('/back-to-index')
@login_required
def back_to_index():
    # Clear session data when going back
    session.pop('keluarga_data', None)
    session.pop('current_member', None)
    session.pop('all_members_data', None)
    session.pop('form_data', None)
    return redirect('/index?clear=true')

@app.route('/lanjutan')
@login_required
def lanjutan():
    keluarga_data = session.get('keluarga_data')
    if not keluarga_data:
        return redirect(url_for('index'))
    return render_template('lanjutan.html', keluarga_data=keluarga_data)

@app.route('/submit-individu', methods=['POST'])
@login_required
def submit_individu():
    try:
        keluarga_data = session.get('keluarga_data')
        if not keluarga_data:
            return jsonify({'success': False, 'message': 'Data keluarga tidak ditemukan'})
        
        current_member = session.get('current_member', 0)
        all_members_data = session.get('all_members_data', [])
        
        # Get form data
        individu_data = {
            'Nama Anggota': request.form.get('nama'),
            'Umur': request.form.get('umur'),
            'Hubungan dengan Kepala Keluarga': request.form.get('hubungan'),
            'Jenis Kelamin': request.form.get('jenis_kelamin'),
            'Status Perkawinan': request.form.get('status_perkawinan'),
            'Pendidikan Terakhir': request.form.get('pendidikan'),
            'Kegiatan Sehari-hari': request.form.get('kegiatan'),
            'Apakah Memiliki Pekerjaan': request.form.get('memiliki_pekerjaan'),
            'Status Pekerjaan yang Diinginkan': request.form.get('status_pekerjaan_diinginkan', ''),
            'Bidang Usaha yang Diminati': request.form.get('bidang_usaha', '')
        }
        
        # Handle "Lainnya" option for bidang usaha
        if individu_data['Bidang Usaha yang Diminati'] == 'Lainnya':
            other_bidang_usaha = request.form.get('other_bidang_usaha_input', '').strip()
            if other_bidang_usaha:
                individu_data['Bidang Usaha yang Diminati'] = other_bidang_usaha
        
        # Add to members list
        all_members_data.append(individu_data)
        current_member += 1
        
        # Update session
        session['all_members_data'] = all_members_data
        session['current_member'] = current_member
        
        remaining = keluarga_data['jumlah_anggota_15plus'] - current_member
        
        if remaining > 0:
            return jsonify({
                'success': True,
                'message': f'Data anggota ke-{current_member} berhasil disimpan',
                'continue_next_member': True,
                'remaining': remaining
            })
        else:
            # All members completed, check if any member has job
            has_working_member = any(member.get('Apakah Memiliki Pekerjaan') == 'Ya' for member in all_members_data)
            
            if has_working_member:
                return jsonify({
                    'success': True,
                    'message': 'Semua data anggota berhasil disimpan. Lanjut ke data pekerjaan.',
                    'redirect': True,
                    'redirect_url': '/pekerjaan'
                })
            else:
                return jsonify({
                    'success': True,
                    'message': 'Semua data anggota berhasil disimpan. Lanjut ke halaman akhir.',
                    'redirect': True,
                    'redirect_url': '/final'
                })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        })

@app.route('/pekerjaan')
@login_required
def pekerjaan():
    keluarga_data = session.get('keluarga_data')
    all_members_data = session.get('all_members_data', [])
    
    if not keluarga_data or not all_members_data:
        return redirect(url_for('index'))
    
    # Filter members who have jobs
    working_members = [member for member in all_members_data if member.get('Apakah Memiliki Pekerjaan') == 'Ya']
    
    if not working_members:
        return redirect(url_for('final_page'))
    
    return render_template('pekerjaan.html', 
                         keluarga_data=keluarga_data, 
                         working_members=working_members)

@app.route('/submit-pekerjaan', methods=['POST'])
@login_required
def submit_pekerjaan():
    try:
        all_members_data = session.get('all_members_data', [])
        if not all_members_data:
            return jsonify({'success': False, 'message': 'Data anggota tidak ditemukan'})
        
        # Get job data from form
        bidang_pekerjaan = request.form.get('bidang_pekerjaan', '')
        lebih_dari_satu_pekerjaan = request.form.get('lebih_dari_satu_pekerjaan', 'Tidak')
        
        # Process main job data
        status_pekerjaan_0 = request.form.get('status_pekerjaan_0', '')
        pemasaran_usaha_0 = request.form.get('pemasaran_usaha_0', '')
        penjualan_marketplace_0 = request.form.get('penjualan_marketplace_0', '')
        
        # Process side jobs if any
        side_job_1_data = {}
        side_job_2_data = {}
        
        if lebih_dari_satu_pekerjaan == 'Ya':
            # Side job 1
            side_job_1_data = {
                'Bidang Pekerjaan Sampingan 1': request.form.get('bidang_pekerjaan_1', ''),
                'Status Pekerjaan Sampingan 1': request.form.get('status_pekerjaan_1', ''),
                'Pemasaran Usaha Sampingan 1': request.form.get('pemasaran_usaha_1', ''),
                'Penjualan Marketplace Sampingan 1': request.form.get('penjualan_marketplace_1', '')
            }
            
            # Side job 2
            side_job_2_data = {
                'Bidang Pekerjaan Sampingan 2': request.form.get('bidang_pekerjaan_2', ''),
                'Status Pekerjaan Sampingan 2': request.form.get('status_pekerjaan_2', ''),
                'Pemasaran Usaha Sampingan 2': request.form.get('pemasaran_usaha_2', ''),
                'Penjualan Marketplace Sampingan 2': request.form.get('penjualan_marketplace_2', '')
            }
        
        # Update the first working member with job data
        for i, member in enumerate(all_members_data):
            if member.get('Apakah Memiliki Pekerjaan') == 'Ya':
                # Add main job data
                member.update({
                    'Bidang Pekerjaan': bidang_pekerjaan,
                    'Status Pekerjaan Utama': status_pekerjaan_0,
                    'Pemasaran Usaha Utama': pemasaran_usaha_0,
                    'Penjualan Marketplace Utama': penjualan_marketplace_0,
                    'Memiliki Lebih dari Satu Pekerjaan': lebih_dari_satu_pekerjaan
                })
                
                # Add side job data if applicable
                if lebih_dari_satu_pekerjaan == 'Ya':
                    member.update(side_job_1_data)
                    member.update(side_job_2_data)
                
                break  # Only update the first working member for now
        
        # Update session
        session['all_members_data'] = all_members_data
        
        return jsonify({
            'success': True,
            'message': 'Data pekerjaan berhasil disimpan',
            'redirect': True,
            'redirect_url': '/final'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        })

@app.route('/final')
@login_required
def final_page():
    keluarga_data = session.get('keluarga_data')
    all_members_data = session.get('all_members_data', [])
    
    if not keluarga_data or not all_members_data:
        return redirect(url_for('index'))
    
    # Add all_members_data to keluarga_data for template
    keluarga_data['all_members_data'] = all_members_data
    
    return render_template('final.html', keluarga_data=keluarga_data)

@app.route('/submit-final', methods=['POST'])
@login_required
def submit_final():
    try:
        keluarga_data = session.get('keluarga_data')
        all_members_data = session.get('all_members_data', [])
        
        if not keluarga_data or not all_members_data:
            return jsonify({'success': False, 'message': 'Data tidak lengkap'})
        
        # Get final form data
        final_data = {
            'nama_pencacah': request.form.get('nama_pencacah'),
            'hp_pencacah': request.form.get('hp_pencacah'),
            'tanggal_pencacah': request.form.get('tanggal_pencacah'),
            'nama_pemberi_jawaban': request.form.get('nama_pemberi_jawaban'),
            'hp_pemberi_jawaban': request.form.get('hp_pemberi_jawaban'),
            'tanggal_pemberi_jawaban': request.form.get('tanggal_pemberi_jawaban'),
            'catatan': request.form.get('catatan', '')
        }
        
        # Save to Excel
        if save_to_excel(keluarga_data, all_members_data, final_data):
            # Clear session data
            session.pop('keluarga_data', None)
            session.pop('all_members_data', None)
            session.pop('current_member', None)
            
            return jsonify({
                'success': True,
                'message': 'Semua data berhasil disimpan!',
                'download_url': '/download/data_sensus.xlsx',
                'redirect_url': '/index?clear=true'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Gagal menyimpan data ke Excel'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        })

@app.route('/edit-keluarga')
@login_required
def edit_keluarga():
    keluarga_data = session.get('keluarga_data')
    if not keluarga_data:
        return redirect(url_for('index'))
    return render_template('edit_keluarga.html', keluarga_data=keluarga_data)

@app.route('/edit-keluarga', methods=['POST'])
@login_required
def update_keluarga():
    try:
        keluarga_data = session.get('keluarga_data', {})
        
        # Update keluarga data
        keluarga_data.update({
            'rt': request.form.get('rt'),
            'rw': request.form.get('rw'),
            'dusun': request.form.get('dusun'),
            'nama_kepala': request.form.get('nama_kepala'),
            'alamat': request.form.get('alamat'),
            'jumlah_anggota': int(request.form.get('jumlah_anggota', 0)),
            'jumlah_anggota_15plus': int(request.form.get('jumlah_anggota_15plus', 0))
        })
        
        session['keluarga_data'] = keluarga_data
        
        return jsonify({
            'success': True,
            'message': 'Data keluarga berhasil diperbarui'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        })

@app.route('/edit-individu/<int:index>')
@login_required
def edit_individu(index):
    keluarga_data = session.get('keluarga_data')
    all_members_data = session.get('all_members_data', [])
    
    if not keluarga_data or not all_members_data or index >= len(all_members_data):
        return redirect(url_for('final_page'))
    
    individu_data = all_members_data[index]
    return render_template('edit_individu.html', 
                         individu_data=individu_data, 
                         index=index,
                         keluarga_data=keluarga_data)

@app.route('/edit-individu/<int:index>', methods=['POST'])
@login_required
def update_individu(index):
    try:
        all_members_data = session.get('all_members_data', [])
        
        if index >= len(all_members_data):
            return jsonify({'success': False, 'message': 'Data anggota tidak ditemukan'})
        
        # Update individual data
        individu_data = {
            'Nama Anggota': request.form.get('nama'),
            'Umur': request.form.get('umur'),
            'Hubungan dengan Kepala Keluarga': request.form.get('hubungan'),
            'Jenis Kelamin': request.form.get('jenis_kelamin'),
            'Status Perkawinan': request.form.get('status_perkawinan'),
            'Pendidikan Terakhir': request.form.get('pendidikan'),
            'Kegiatan Sehari-hari': request.form.get('kegiatan'),
            'Apakah Memiliki Pekerjaan': request.form.get('memiliki_pekerjaan'),
            'Status Pekerjaan yang Diinginkan': request.form.get('status_pekerjaan_diinginkan', ''),
            'Bidang Usaha yang Diminati': request.form.get('bidang_usaha', '')
        }
        
        # Handle "Lainnya" option for bidang usaha
        if individu_data['Bidang Usaha yang Diminati'] == 'Lainnya':
            other_bidang_usaha = request.form.get('other_bidang_usaha_input', '').strip()
            if other_bidang_usaha:
                individu_data['Bidang Usaha yang Diminati'] = other_bidang_usaha
        
        # Handle job data if applicable
        if individu_data['Apakah Memiliki Pekerjaan'] == 'Ya':
            # Main job data
            individu_data.update({
                'Bidang Pekerjaan': request.form.get('bidang_pekerjaan', ''),
                'Status Pekerjaan Utama': request.form.get('status_pekerjaan_0', ''),
                'Pemasaran Usaha Utama': request.form.get('pemasaran_usaha_0', ''),
                'Penjualan Marketplace Utama': request.form.get('penjualan_marketplace_0', ''),
                'Memiliki Lebih dari Satu Pekerjaan': request.form.get('lebih_dari_satu_pekerjaan', 'Tidak')
            })
            
            # Side job data if applicable
            if request.form.get('lebih_dari_satu_pekerjaan') == 'Ya':
                individu_data.update({
                    'Bidang Pekerjaan Sampingan 1': request.form.get('bidang_pekerjaan_1', ''),
                    'Status Pekerjaan Sampingan 1': request.form.get('status_pekerjaan_1', ''),
                    'Pemasaran Usaha Sampingan 1': request.form.get('pemasaran_usaha_1', ''),
                    'Penjualan Marketplace Sampingan 1': request.form.get('penjualan_marketplace_1', ''),
                    'Bidang Pekerjaan Sampingan 2': request.form.get('bidang_pekerjaan_2', ''),
                    'Status Pekerjaan Sampingan 2': request.form.get('status_pekerjaan_2', ''),
                    'Pemasaran Usaha Sampingan 2': request.form.get('pemasaran_usaha_2', ''),
                    'Penjualan Marketplace Sampingan 2': request.form.get('penjualan_marketplace_2', '')
                })
        
        # Update the member data
        all_members_data[index] = individu_data
        session['all_members_data'] = all_members_data
        
        return jsonify({
            'success': True,
            'message': 'Data anggota berhasil diperbarui'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        })

@app.route('/download/<filename>')
@login_required
def download_file(filename):
    try:
        if filename == 'data_sensus.xlsx' and os.path.exists(EXCEL_FILE):
            return send_file(EXCEL_FILE, as_attachment=True, download_name=filename)
        else:
            return jsonify({'error': 'File tidak ditemukan'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/check-file')
@login_required
def check_file():
    exists = os.path.exists(EXCEL_FILE)
    return jsonify({'exists': exists})

if __name__ == '__main__':
    app.run(debug=True)
